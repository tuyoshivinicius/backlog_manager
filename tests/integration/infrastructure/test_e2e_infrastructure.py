"""Testes E2E da infraestrutura completa."""
import pytest
from openpyxl import Workbook

from backlog_manager.infrastructure.database.sqlite_connection import SQLiteConnection
from backlog_manager.infrastructure.database.unit_of_work import UnitOfWork
from backlog_manager.infrastructure.excel.openpyxl_excel_service import OpenpyxlExcelService


def test_full_flow_excel_import_to_database_to_export(tmp_path):
    """
    Teste E2E completo: Import Excel → Salvar DB → Recuperar DB → Export Excel

    Fluxo:
    1. Criar arquivo Excel com histórias
    2. Importar histórias do Excel
    3. Salvar histórias no banco SQLite
    4. Recuperar histórias do banco
    5. Exportar histórias para novo arquivo Excel
    6. Validar que tudo foi preservado
    """
    # Resetar singleton
    SQLiteConnection._instance = None
    SQLiteConnection._connection = None

    db_path = tmp_path / "test.db"
    import_file = tmp_path / "import.xlsx"
    export_file = tmp_path / "export.xlsx"

    # 1. Criar Excel de importação
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Component", "Nome", "StoryPoint", "Deps"])
    ws.append(["US-001", "Autenticação", "Login", 5, ""])
    ws.append(["US-002", "Autenticação", "Logout", 3, ""])
    ws.append(["US-003", "Dashboard", "Exibir métricas", 8, ""])
    wb.save(import_file)

    # 2. Importar histórias
    excel_service = OpenpyxlExcelService()
    stories_dto, stats, columns_present = excel_service.import_stories(str(import_file))

    assert len(stories_dto) == 3
    assert stories_dto[0].id == "US-001"
    assert stories_dto[1].id == "US-002"
    assert stories_dto[2].id == "US-003"

    # Converter DTOs para entities
    from backlog_manager.application.dto.converters import dto_to_story
    stories = [dto_to_story(dto) for dto in stories_dto]

    # 3. Salvar no banco
    with UnitOfWork(str(db_path)) as uow:
        for story in stories:
            uow.stories.save(story)
        uow.commit()

    # Resetar singleton para nova conexão
    SQLiteConnection._instance = None
    SQLiteConnection._connection = None

    # 4. Recuperar do banco
    with UnitOfWork(str(db_path)) as uow:
        recovered_stories = uow.stories.find_all()

    assert len(recovered_stories) == 3
    assert recovered_stories[0].id == "US-001"
    assert recovered_stories[0].component == "Autenticação"
    assert recovered_stories[0].name == "Login"
    assert recovered_stories[0].story_point.value == 5

    # 5. Exportar para Excel
    from backlog_manager.application.dto.converters import story_to_dto
    recovered_stories_dto = [story_to_dto(story) for story in recovered_stories]
    excel_service.export_backlog(str(export_file), recovered_stories_dto)

    assert export_file.exists()

    # 6. Validar Excel exportado
    from openpyxl import load_workbook

    wb_export = load_workbook(export_file)
    ws_export = wb_export.active

    # Validar cabeçalho
    assert ws_export.cell(1, 1).value == "Prioridade"
    assert ws_export.cell(1, 2).value == "ID"
    assert ws_export.cell(1, 3).value == "Component"

    # Validar dados
    assert ws_export.cell(2, 2).value == "US-001"
    assert ws_export.cell(2, 3).value == "Autenticação"
    assert ws_export.cell(2, 4).value == "Login"
    assert ws_export.cell(2, 8).value == 5  # Story Point

    assert ws_export.cell(3, 2).value == "US-002"
    assert ws_export.cell(4, 2).value == "US-003"


def test_transaction_rollback_on_error(tmp_path):
    """Deve fazer rollback se erro ocorrer durante transação."""
    # Resetar singleton
    SQLiteConnection._instance = None
    SQLiteConnection._connection = None

    db_path = tmp_path / "test.db"

    from backlog_manager.domain.entities.story import Story
    from backlog_manager.domain.value_objects.story_point import StoryPoint
    from backlog_manager.domain.value_objects.story_status import StoryStatus

    story1 = Story(
        id="US-001",
        component="F1",
        name="S1",
        status=StoryStatus.BACKLOG,
        priority=0,
        developer_id=None,
        dependencies=[],
        story_point=StoryPoint(5),
    )

    # Tentar salvar mas simular erro
    try:
        with UnitOfWork(str(db_path)) as uow:
            uow.stories.save(story1)
            # Simular erro antes do commit
            raise RuntimeError("Erro simulado")
    except RuntimeError:
        pass

    # Resetar singleton
    SQLiteConnection._instance = None
    SQLiteConnection._connection = None

    # Verificar que rollback foi feito
    with UnitOfWork(str(db_path)) as uow:
        stories = uow.stories.find_all()

    assert len(stories) == 0  # Rollback descartou a história


def test_developer_and_story_relationship(tmp_path):
    """Deve gerenciar relacionamento entre desenvolvedor e histórias."""
    # Resetar singleton
    SQLiteConnection._instance = None
    SQLiteConnection._connection = None

    db_path = tmp_path / "test.db"

    from backlog_manager.domain.entities.developer import Developer
    from backlog_manager.domain.entities.story import Story
    from backlog_manager.domain.value_objects.story_point import StoryPoint
    from backlog_manager.domain.value_objects.story_status import StoryStatus

    # Criar desenvolvedor
    developer = Developer(id="DEV-001", name="Alice")

    # Criar história alocada ao desenvolvedor
    story = Story(
        id="US-001",
        component="F1",
        name="S1",
        status=StoryStatus.BACKLOG,
        priority=0,
        developer_id="DEV-001",
        dependencies=[],
        story_point=StoryPoint(5),
    )

    # Salvar no banco
    with UnitOfWork(str(db_path)) as uow:
        uow.developers.save(developer)
        uow.stories.save(story)
        uow.commit()

    # Resetar singleton
    SQLiteConnection._instance = None
    SQLiteConnection._connection = None

    # Recuperar
    with UnitOfWork(str(db_path)) as uow:
        found_dev = uow.developers.find_by_id("DEV-001")
        found_story = uow.stories.find_by_id("US-001")

    assert found_dev.name == "Alice"
    assert found_story.developer_id == "DEV-001"

    # Resetar singleton
    SQLiteConnection._instance = None
    SQLiteConnection._connection = None

    # Deletar desenvolvedor (deve setar NULL na história devido a ON DELETE SET NULL)
    with UnitOfWork(str(db_path)) as uow:
        uow.developers.delete("DEV-001")
        uow.commit()

    # Resetar singleton
    SQLiteConnection._instance = None
    SQLiteConnection._connection = None

    # Verificar que história ainda existe mas sem desenvolvedor
    with UnitOfWork(str(db_path)) as uow:
        found_story = uow.stories.find_by_id("US-001")

    assert found_story is not None
    assert found_story.developer_id is None  # Foreign key foi setada para NULL
