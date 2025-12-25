"""Testes de integração para OpenpyxlExcelService."""
import pytest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from backlog_manager.domain.entities.story import Story
from backlog_manager.domain.value_objects.story_point import StoryPoint
from backlog_manager.domain.value_objects.story_status import StoryStatus
from backlog_manager.infrastructure.excel.openpyxl_excel_service import OpenpyxlExcelService


@pytest.fixture
def excel_service():
    return OpenpyxlExcelService()


@pytest.fixture
def valid_excel_file(tmp_path):
    """Cria arquivo Excel válido para testes."""
    file_path = tmp_path / "backlog.xlsx"

    wb = Workbook()
    ws = wb.active

    # Cabeçalho (NOVO: 5 colunas)
    ws.append(["ID", "Component", "Nome", "StoryPoint", "Deps"])

    # Dados
    ws.append(["US-001", "Autenticação", "Login de usuário", 5, ""])
    ws.append(["US-002", "Autenticação", "Logout", 3, ""])
    ws.append(["US-003", "Dashboard", "Exibir métricas", 8, ""])

    wb.save(file_path)
    return str(file_path)


def test_import_valid_excel(excel_service, valid_excel_file):
    """Deve importar histórias de Excel válido."""
    stories, stats, columns_present = excel_service.import_stories(valid_excel_file)

    assert len(stories) == 3
    assert stories[0].id == "US-001"
    assert stories[0].component == "Autenticação"
    assert stories[0].name == "Login de usuário"
    assert stories[0].story_point == 5
    assert stories[1].id == "US-002"
    assert stories[2].id == "US-003"

    # Verificar estatísticas
    assert stats["total_importadas"] == 3
    assert stats["ignoradas_duplicadas"] == 0
    assert stats["ignoradas_invalidas"] == 0

    # Verificar colunas presentes (formato legado)
    assert "id" in columns_present
    assert "component" in columns_present
    assert "nome" in columns_present
    assert "story_point" in columns_present
    assert "deps" in columns_present


def test_import_file_not_found(excel_service):
    """Deve lançar erro se arquivo não existe."""
    with pytest.raises(FileNotFoundError, match="Arquivo não encontrado"):
        excel_service.import_stories("arquivo_inexistente.xlsx")


def test_import_invalid_header(excel_service, tmp_path):
    """Deve lançar erro se cabeçalho inválido (colunas obrigatórias ausentes)."""
    file_path = tmp_path / "invalid_header.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.append(["Coluna1", "Coluna2", "Coluna3", "Coluna4", "Coluna5"])  # Cabeçalho errado
    ws.append(["ID1", "Component1", "Story1", 5, ""])
    wb.save(file_path)

    with pytest.raises(ValueError, match="Colunas obrigatórias ausentes"):
        excel_service.import_stories(str(file_path))


def test_import_invalid_story_point_raises_error(excel_service, tmp_path):
    """Deve ignorar linha com Story Point inválido e adicionar warning."""
    file_path = tmp_path / "invalid.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Component", "Nome", "StoryPoint", "Deps"])
    ws.append(["US-001", "Component1", "Story1", 7, ""])  # 7 é inválido
    ws.append(["US-002", "Component2", "Story2", 5, ""])  # Válido
    wb.save(file_path)

    stories, stats, columns_present = excel_service.import_stories(str(file_path))

    # História com story point inválido deve ser ignorada
    assert len(stories) == 1
    assert stories[0].id == "US-002"
    assert stats["ignoradas_invalidas"] == 1
    assert any("Story Point inválido" in w for w in stats["warnings"])


def test_import_missing_data(excel_service, tmp_path):
    """Deve ignorar linha com dados obrigatórios faltando."""
    file_path = tmp_path / "missing_data.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Component", "Nome", "StoryPoint", "Deps"])
    ws.append(["US-001", "Component1", None, 5, ""])  # Nome faltando
    ws.append(["US-002", "Component2", "Story2", 5, ""])  # Válido
    wb.save(file_path)

    stories, stats, columns_present = excel_service.import_stories(str(file_path))

    # Linha com nome vazio deve ser ignorada
    assert len(stories) == 1
    assert stories[0].id == "US-002"
    assert stats["ignoradas_invalidas"] == 1
    assert any("Nome vazio" in w for w in stats["warnings"])


def test_export_creates_formatted_excel(excel_service, tmp_path):
    """Deve exportar backlog para Excel formatado."""
    from backlog_manager.application.dto.story_dto import StoryDTO

    file_path = tmp_path / "export.xlsx"

    stories = [
        StoryDTO(
            id="US-001",
            component="F1",
            name="S1",
            feature_id="feature_default",
            status="BACKLOG",
            priority=0,
            developer_id="DEV-001",
            dependencies=["US-002"],
            story_point=5,
            start_date=None,
            end_date=None,
            duration=None,
        )
    ]

    excel_service.export_backlog(str(file_path), stories)

    assert Path(file_path).exists()

    # Verificar conteúdo
    wb = load_workbook(file_path)
    ws = wb.active

    # Cabeçalho (nova ordem: Prioridade, Feature, Onda, ID, Component, ...)
    assert ws.cell(1, 1).value == "Prioridade"
    assert ws.cell(1, 2).value == "Feature"
    assert ws.cell(1, 3).value == "Onda"
    assert ws.cell(1, 4).value == "ID"
    assert ws.cell(1, 5).value == "Component"

    # Dados (colunas deslocadas +2 por Feature e Onda)
    assert ws.cell(2, 1).value == 0  # priority
    assert ws.cell(2, 4).value == "US-001"  # id (coluna 4)
    assert ws.cell(2, 5).value == "F1"  # component (coluna 5)
    assert ws.cell(2, 6).value == "S1"  # name (coluna 6)
    assert ws.cell(2, 7).value == "BACKLOG"  # status (coluna 7)
    assert ws.cell(2, 8).value == "DEV-001"  # developer_id (coluna 8)
    assert ws.cell(2, 9).value == "US-002"  # dependencies (coluna 9)
    assert ws.cell(2, 10).value == 5  # story_point (coluna 10)


def test_export_multiple_stories(excel_service, tmp_path):
    """Deve exportar múltiplas histórias."""
    from backlog_manager.application.dto.story_dto import StoryDTO

    file_path = tmp_path / "export_multiple.xlsx"

    stories = [
        StoryDTO(
            id="US-001",
            component="F1",
            name="S1",
            feature_id="feature_default",
            status="BACKLOG",
            priority=0,
            developer_id=None,
            dependencies=[],
            story_point=3,
            start_date=None,
            end_date=None,
            duration=None,
        ),
        StoryDTO(
            id="US-002",
            component="F1",
            name="S2",
            feature_id="feature_default",
            status="EXECUCAO",
            priority=1,
            developer_id="DEV-001",
            dependencies=["US-001"],
            story_point=5,
            start_date=None,
            end_date=None,
            duration=None,
        ),
        StoryDTO(
            id="US-003",
            component="F2",
            name="S3",
            feature_id="feature_default",
            status="CONCLUIDO",
            priority=2,
            developer_id="DEV-002",
            dependencies=[],
            story_point=8,
            start_date=None,
            end_date=None,
            duration=None,
        ),
    ]

    excel_service.export_backlog(str(file_path), stories)

    # Verificar
    wb = load_workbook(file_path)
    ws = wb.active

    # Deve ter cabeçalho + 3 linhas de dados
    assert ws.max_row == 4
    # ID agora está na coluna 4 (após Prioridade, Feature, Onda)
    assert ws.cell(2, 4).value == "US-001"
    assert ws.cell(3, 4).value == "US-002"
    assert ws.cell(4, 4).value == "US-003"
