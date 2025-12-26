"""Implementação do Excel Service usando openpyxl."""
import logging
from datetime import date
from pathlib import Path
from typing import Any, List, Tuple, Set, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backlog_manager.application.interfaces.services.excel_service import ExcelService
from backlog_manager.application.dto.story_dto import StoryDTO
from backlog_manager.domain.entities.story import Story
from backlog_manager.domain.value_objects.story_point import StoryPoint
from backlog_manager.domain.value_objects.story_status import StoryStatus

logger = logging.getLogger(__name__)


class OpenpyxlExcelService(ExcelService):
    """
    Implementação do serviço de Excel usando openpyxl.

    Responsabilidades:
    - Importar histórias de planilha Excel (5 colunas: ID, Component, Nome, StoryPoint, Deps)
    - Exportar backlog para Excel com formatação
    - Validar formato e dados
    - Detectar e ignorar IDs duplicados
    - Processar dependências
    """

    # Colunas esperadas no import (NOVO: 5 colunas)
    IMPORT_COLUMNS = ["ID", "Component", "Nome", "StoryPoint", "Deps"]

    # Colunas do export
    EXPORT_COLUMNS = [
        "Prioridade",
        "Feature",
        "Onda",
        "ID",
        "Component",
        "Nome",
        "Status",
        "Desenvolvedor",
        "Dependências",
        "SP",
        "Início",
        "Fim",
        "Duração",
    ]

    # Mapeamento flexível de colunas (case-insensitive)
    # Mapeia campo normalizado → aliases aceitos
    COLUMN_ALIASES = {
        "id": ["id"],
        "component": ["component"],
        "nome": ["nome", "name"],
        "story_point": ["storypoint", "sp"],
        "deps": ["deps", "dependencias", "dependências"],
        "status": ["status"],
        "desenvolvedor": ["desenvolvedor", "developer", "developer_id"],
        "prioridade": ["prioridade", "priority"],
        "feature": ["feature"],
        "onda": ["onda", "wave"],
    }

    def _detect_and_map_columns(self, header: List[Any]) -> dict[str, int]:
        """
        Detecta formato da planilha e mapeia colunas (case-insensitive).

        Args:
            header: Lista de nomes de colunas do cabeçalho Excel

        Returns:
            Dict mapeando campo normalizado → índice da coluna (0-based)
            Ex: {"id": 1, "component": 2, "story_point": 7}

        Raises:
            ValueError: Se colunas obrigatórias ausentes
        """
        # Normalizar cabeçalho (lowercase, strip)
        normalized_header = []
        for col in header:
            if col is None:
                normalized_header.append("")
            else:
                normalized_header.append(str(col).strip().lower())

        # Mapear cada campo normalizado → índice da coluna
        column_map: dict[str, int] = {}

        for field, aliases in self.COLUMN_ALIASES.items():
            for col_idx, col_name in enumerate(normalized_header):
                if col_name in [alias.lower() for alias in aliases]:
                    column_map[field] = col_idx
                    break

        # Validar colunas obrigatórias
        required_fields = ["id", "component", "nome"]
        missing_fields = [field for field in required_fields if field not in column_map]

        if missing_fields:
            missing_names = [field.title() for field in missing_fields]
            raise ValueError(
                f"Colunas obrigatórias ausentes: {', '.join(missing_names)}"
            )

        # Validar presença de story_point (aceita "SP" ou "StoryPoint")
        if "story_point" not in column_map:
            raise ValueError(
                "Nenhuma coluna de Story Point encontrada. Use 'SP' ou 'StoryPoint'"
            )

        return column_map

    def _extract_value(self, row: Tuple, column_map: dict[str, int], field: str) -> Any:
        """
        Extrai valor de campo usando mapeamento flexível.

        Args:
            row: Tupla de valores da linha Excel
            column_map: Mapeamento campo → índice
            field: Nome do campo normalizado

        Returns:
            Valor do campo ou None se campo não presente
        """
        if field not in column_map:
            return None

        col_idx = column_map[field]
        if col_idx >= len(row):
            return None

        return row[col_idx]

    def _get_present_columns(self, column_map: dict[str, int]) -> Set[str]:
        """
        Retorna set de campos normalizados presentes na planilha.

        Args:
            column_map: Mapeamento campo → índice

        Returns:
            Set de campos presentes
            Ex: {"id", "component", "nome", "story_point", "deps", "status"}
        """
        return set(column_map.keys())

    def import_stories(
        self,
        filepath: str,
        existing_ids: Optional[Set[str]] = None
    ) -> Tuple[List[StoryDTO], dict, Set[str]]:
        """
        Importa histórias de arquivo Excel com mapeamento flexível de colunas.

        Suporta dois formatos:
        - Legado (5 colunas): ID, Component, Nome, StoryPoint, Deps
        - Completo (11 colunas): Prioridade, ID, Component, Nome, Status, Desenvolvedor, Dependências, SP, Início, Fim, Duração

        Mapeamento case-insensitive de aliases: StoryPoint/SP, Deps/Dependências, etc.

        Args:
            filepath: Caminho do arquivo .xlsx
            existing_ids: IDs de histórias já existentes no banco

        Returns:
            Tupla (stories_dto, stats, columns_present) onde:
            - stories_dto: Lista de histórias válidas importadas
            - stats: Estatísticas da importação
            - columns_present: Set de campos presentes na planilha

        Raises:
            FileNotFoundError: Se arquivo não existe
            ValueError: Se colunas obrigatórias ausentes
        """
        logger.info(f"Iniciando importação de Excel: '{filepath}'")

        if existing_ids is None:
            existing_ids = set()

        logger.debug(f"IDs existentes no banco: {len(existing_ids)}")

        if not Path(filepath).exists():
            logger.error(f"Arquivo não encontrado: '{filepath}'")
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

        try:
            workbook = load_workbook(filepath)
            sheet = workbook.active
            logger.debug(f"Arquivo Excel carregado com sucesso: '{filepath}'")

            # Detectar e mapear colunas (case-insensitive, flexível)
            header = [cell.value for cell in sheet[1]]
            column_map = self._detect_and_map_columns(header)
            columns_present = self._get_present_columns(column_map)
            logger.debug(f"Colunas detectadas: {sorted(columns_present)}")

            # Inicializar estatísticas
            stats: dict[str, Any] = {
                "total_processadas": 0,
                "total_importadas": 0,
                "ignoradas_duplicadas": 0,
                "ignoradas_invalidas": 0,
                "deps_ignoradas": 0,
                "warnings": []
            }

            temp_stories = []  # Lista temporária: (story_id, story_dto, row_num, deps_value)
            id_counts: dict[str, int] = {}     # Mapear ID → contagem de ocorrências
            generated_id_counter = 1

            # FASE 1: Processar todas as linhas e detectar duplicatas
            logger.debug("FASE 1: Processando linhas e detectando duplicatas")
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                stats["total_processadas"] += 1

                # Extrair valores usando mapeamento flexível
                id_value = self._extract_value(row, column_map, "id")
                component = self._extract_value(row, column_map, "component")
                name = self._extract_value(row, column_map, "nome")
                sp_value = self._extract_value(row, column_map, "story_point")
                deps_value = self._extract_value(row, column_map, "deps")
                status = self._extract_value(row, column_map, "status")
                desenvolvedor = self._extract_value(row, column_map, "desenvolvedor")
                prioridade = self._extract_value(row, column_map, "prioridade")

                # Validar campos obrigatórios (Component e Nome)
                if not component or str(component).strip() == "":
                    stats["ignoradas_invalidas"] += 1
                    stats["warnings"].append(f"Linha {row_num}: Component vazio - linha ignorada")
                    continue

                if not name or str(name).strip() == "":
                    stats["ignoradas_invalidas"] += 1
                    stats["warnings"].append(f"Linha {row_num}: Nome vazio - linha ignorada")
                    continue

                # Gerar ID se vazio
                if id_value and str(id_value).strip():
                    story_id = str(id_value).strip()
                else:
                    story_id = f"US-{generated_id_counter:03d}"
                    generated_id_counter += 1

                # Contar ocorrências de ID (para detectar duplicatas)
                id_counts[story_id] = id_counts.get(story_id, 0) + 1

                # Validar Story Point (opcional)
                sp_validated = None
                if sp_value is not None and str(sp_value).strip() != "":
                    try:
                        sp_validated = int(sp_value)
                        if sp_validated not in [3, 5, 8, 13]:
                            stats["ignoradas_invalidas"] += 1
                            stats["warnings"].append(
                                f"Linha {row_num}: Story Point inválido ({sp_value}) - linha ignorada"
                            )
                            continue
                    except (ValueError, TypeError):
                        stats["ignoradas_invalidas"] += 1
                        stats["warnings"].append(
                            f"Linha {row_num}: Story Point inválido ({sp_value}) - linha ignorada"
                        )
                        continue

                # Processar status (com fallback para BACKLOG)
                status_str = "BACKLOG"
                if status and str(status).strip():
                    status_str = str(status).strip().upper()
                    # Validar se status é válido
                    valid_statuses = ["BACKLOG", "EXECUCAO", "TESTES", "CONCLUIDO", "IMPEDIDO"]
                    if status_str not in valid_statuses:
                        stats["warnings"].append(
                            f"Linha {row_num}: Status inválido '{status}', usando 'BACKLOG'"
                        )
                        status_str = "BACKLOG"

                # Processar desenvolvedor (opcional)
                developer_id = None
                if desenvolvedor and str(desenvolvedor).strip():
                    developer_id = str(desenvolvedor).strip()

                # Processar prioridade (com fallback para 0)
                priority_value = 0
                if prioridade is not None:
                    try:
                        priority_value = int(prioridade)
                        if priority_value < 0:
                            priority_value = 0
                    except (ValueError, TypeError):
                        priority_value = 0

                # NOVO: Extrair Feature e Onda
                feature_value = self._extract_value(row, column_map, "feature")
                onda_value = self._extract_value(row, column_map, "onda")

                # Processar feature_name
                feature_name = None
                if feature_value and str(feature_value).strip():
                    feature_name = str(feature_value).strip()

                # Processar wave
                wave = None
                if onda_value is not None:
                    try:
                        wave = int(onda_value)
                        if wave <= 0:
                            wave = None
                    except (ValueError, TypeError):
                        wave = None

                # Criar DTO temporário (dependências serão processadas depois)
                story_dto = StoryDTO(
                    id=story_id,
                    component=str(component).strip(),
                    name=str(name).strip(),
                    feature_id=None,  # Será preenchido pelo UseCase (None = sem feature)
                    status=status_str,
                    priority=priority_value,
                    developer_id=developer_id,
                    dependencies=[],  # Processado na FASE 3
                    story_point=sp_validated,
                    start_date=None,
                    end_date=None,
                    duration=None,
                    feature_name=feature_name,  # NOVO
                    wave=wave,  # NOVO
                )

                temp_stories.append((story_id, story_dto, row_num, deps_value))

            logger.debug(f"FASE 1 concluída: {len(temp_stories)} histórias processadas")

            # FASE 2: Identificar IDs duplicados
            logger.debug("FASE 2: Identificando IDs duplicados")
            duplicated_ids = {story_id for story_id, count in id_counts.items() if count > 1}
            if duplicated_ids:
                logger.warning(f"IDs duplicados encontrados: {len(duplicated_ids)} IDs afetam múltiplas linhas")

            # Registrar warnings para IDs duplicados
            for story_id in duplicated_ids:
                count = id_counts[story_id]
                stats["ignoradas_duplicadas"] += count
                for i in range(count):
                    stats["warnings"].append(
                        f"ID '{story_id}' duplicado na planilha - {count} linhas ignoradas"
                    )

            # FASE 3: Filtrar histórias válidas (sem duplicatas) e processar dependências
            logger.debug("FASE 3: Processando dependências")
            valid_stories = []
            all_story_ids = {sid for sid, _, _, _ in temp_stories if sid not in duplicated_ids}

            for story_id, story_dto, row_num, deps_value in temp_stories:
                # Ignorar histórias com ID duplicado
                if story_id in duplicated_ids:
                    continue

                # Processar dependências
                valid_deps, invalid_deps = self._process_dependencies(
                    deps_value, all_story_ids, existing_ids, story_id
                )

                story_dto.dependencies = valid_deps

                # Registrar dependências inválidas
                for invalid_dep in invalid_deps:
                    stats["deps_ignoradas"] += 1
                    stats["warnings"].append(
                        f"Linha {row_num}: Dependência '{invalid_dep}' não encontrada - removida da história '{story_id}'"
                    )

                valid_stories.append(story_dto)

            logger.debug(f"FASE 3 concluída: {len(valid_stories)} histórias válidas")

            # FASE 4: Ajustar prioridades sequencialmente (se coluna prioridade NÃO presente)
            if "prioridade" not in columns_present:
                logger.debug("FASE 4: Ajustando prioridades sequencialmente (coluna 'prioridade' ausente)")
                for idx, story_dto in enumerate(valid_stories, start=1):
                    story_dto.priority = idx

            stats["total_importadas"] = len(valid_stories)

            logger.info(
                f"Importação concluída: {stats['total_importadas']} histórias importadas, "
                f"{stats['ignoradas_duplicadas']} duplicadas, "
                f"{stats['ignoradas_invalidas']} inválidas, "
                f"{stats['deps_ignoradas']} dependências removidas"
            )

            return valid_stories, stats, columns_present

        except FileNotFoundError:
            raise
        except ValueError as e:
            logger.error(f"Erro de validação ao importar Excel: {e}")
            raise
        except Exception as e:
            logger.error(f"Erro ao importar arquivo Excel '{filepath}': {e}", exc_info=True)
            raise

    def _process_dependencies(
        self,
        deps_value: Any,
        all_story_ids: Set[str],
        existing_ids: Set[str],
        story_id: str
    ) -> Tuple[List[str], List[str]]:
        """
        Processa campo Deps e valida dependências.

        Args:
            deps_value: Valor do campo Deps (string com vírgulas ou None)
            all_story_ids: IDs de todas as histórias na planilha
            existing_ids: IDs de histórias já existentes no banco
            story_id: ID da história corrente (não pode depender de si mesma)

        Returns:
            Tupla (valid_deps, invalid_deps)
        """
        if not deps_value or str(deps_value).strip() == "":
            return [], []

        # Split por vírgula e limpar espaços
        deps = [dep.strip() for dep in str(deps_value).split(',') if dep.strip()]

        valid_deps = []
        invalid_deps = []

        for dep_id in deps:
            # Não pode depender de si mesma
            if dep_id == story_id:
                invalid_deps.append(dep_id)
                continue

            # Validar se dependência existe (na planilha ou no banco)
            if dep_id in all_story_ids or dep_id in existing_ids:
                valid_deps.append(dep_id)
            else:
                invalid_deps.append(dep_id)

        return valid_deps, invalid_deps

    def export_backlog(self, filepath: str, stories: List[StoryDTO]) -> None:
        """
        Exporta histórias para arquivo Excel.

        Args:
            filepath: Caminho do arquivo .xlsx a criar
            stories: Lista de histórias a exportar (DTOs)
        """
        logger.info(f"Iniciando exportação de {len(stories)} histórias para Excel: '{filepath}'")

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Backlog"
            logger.debug("Workbook criado")

            # Escrever cabeçalho
            for col_num, column_name in enumerate(self.EXPORT_COLUMNS, start=1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = column_name
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Escrever dados
            for row_num, story in enumerate(stories, start=2):
                sheet.cell(row=row_num, column=1).value = story.priority
                sheet.cell(row=row_num, column=2).value = story.feature_name or ""
                sheet.cell(row=row_num, column=3).value = story.wave or ""
                sheet.cell(row=row_num, column=4).value = story.id
                sheet.cell(row=row_num, column=5).value = story.component
                sheet.cell(row=row_num, column=6).value = story.name
                sheet.cell(row=row_num, column=7).value = story.status
                sheet.cell(row=row_num, column=8).value = story.developer_id or ""
                sheet.cell(row=row_num, column=9).value = ", ".join(story.dependencies) if story.dependencies else ""
                sheet.cell(row=row_num, column=10).value = story.story_point
                sheet.cell(row=row_num, column=11).value = story.start_date.strftime("%d/%m/%Y") if story.start_date else ""
                sheet.cell(row=row_num, column=12).value = story.end_date.strftime("%d/%m/%Y") if story.end_date else ""
                sheet.cell(row=row_num, column=13).value = story.duration or ""

            # Auto-ajustar largura das colunas
            for col_num in range(1, len(self.EXPORT_COLUMNS) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0

                for cell in sheet[column_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Adicionar bordas
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
            )

            for row in sheet.iter_rows(min_row=1, max_row=len(stories) + 1):
                for cell in row:
                    cell.border = thin_border

            # Salvar arquivo
            workbook.save(filepath)
            logger.info(f"Exportação concluída com sucesso: {len(stories)} histórias salvas em '{filepath}'")

        except Exception as e:
            logger.error(f"Erro ao exportar backlog para Excel '{filepath}': {e}", exc_info=True)
            raise
